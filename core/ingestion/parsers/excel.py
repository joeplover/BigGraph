from pathlib import Path
from openpyxl import load_workbook
from core.ingestion.parser_base import ParsedDocument, ParsedSection


class ExcelParser:
    supported_extensions = {".xlsx", ".xlsm"}
    def parse(self, path: Path) -> ParsedDocument:
        wb = load_workbook(str(path), data_only=True, read_only=True)
        sections, md_parts = [], []
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            cleaned = [["" if v is None else str(v).strip() for v in row] for row in rows if any(v is not None and str(v).strip() for v in row)]
            if not cleaned:
                continue
            cols = max(len(r) for r in cleaned)
            normalized = [r + [""] * (cols - len(r)) for r in cleaned]
            tbl = f"### Sheet: {sheet.title}\n"
            tbl += "| " + " | ".join(normalized[0]) + " |\n"
            tbl += "| " + " | ".join(["---"] * cols) + " |\n"
            for row in normalized[1:]:
                tbl += "| " + " | ".join(row) + " |\n"
            md_parts.append(tbl)
            sections.append(ParsedSection(heading_path=f"{path.stem} / {sheet.title}", text=tbl, content_type="table", metadata={"sheet_name": sheet.title, "rows": len(normalized) - 1}))
        md_text = "\n\n".join(md_parts)
        return ParsedDocument(title=path.stem, source_path=str(path), raw_text=md_text, normalized_markdown=md_text, sections=sections, metadata={"parser": "excel", "sheets": len(wb.sheetnames)})
